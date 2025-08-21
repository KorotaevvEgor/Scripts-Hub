from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth import logout
from django.contrib import messages
from django.http import JsonResponse
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.core.paginator import Paginator
from django.db.models import Q
import zoneinfo
from .models import Script, ScriptRun, Vacancy, VacancyRun
from .parser import HHVacancyParserDjango
import json
import threading
import datetime


def get_accessible_scripts_for_user(user, is_active=True):
    """
    Возвращает queryset скриптов, к которым у пользователя есть доступ
    """
    if not user.is_authenticated:
        return Script.objects.none()
    
    if user.is_superuser:
        # Суперпользователь видит все скрипты
        if is_active is None:
            return Script.objects.all()
        return Script.objects.filter(is_active=is_active)
    
    # Обычный пользователь видит:
    # 1. Созданные им скрипты
    # 2. Скрипты, в которых он указан в allowed_users
    query = Q(created_by=user) | Q(allowed_users=user)
    
    if is_active is None:
        return Script.objects.filter(query).distinct()
    
    return Script.objects.filter(query, is_active=is_active).distinct()


def get_user_scripts_context(request):
    """Контекст-процессор для передачи скриптов пользователя в шаблоны"""
    if request.user.is_authenticated:
        user_scripts = get_accessible_scripts_for_user(request.user, is_active=True).order_by('name')[:10]
        return {'user_scripts_menu': user_scripts}
    return {}


@require_http_methods(["GET", "POST"])
def logout_view(request):
    """Выход из системы"""
    logout(request)
    messages.success(request, 'Вы успешно вышли из системы.')
    return redirect('login')


@login_required
def home_view(request):
    """Главная страница"""
    user_scripts = get_accessible_scripts_for_user(request.user, is_active=True)
    recent_runs = ScriptRun.objects.filter(
        script__in=user_scripts
    ).order_by('-started_at')[:5]
    
    # Получаем запуски за сегодня
    # Используем локальную дату для определения сегодняшнего дня
    today_local = timezone.localtime().date()
    
    # Создаем начало и конец дня в локальном часовом поясе
    current_tz = timezone.get_current_timezone()
    start_of_day_local = datetime.datetime.combine(today_local, datetime.time.min, tzinfo=current_tz)
    end_of_day_local = datetime.datetime.combine(today_local, datetime.time.max, tzinfo=current_tz)
    
    # Преобразуем в UTC для сравнения с базой данных
    utc_tz = zoneinfo.ZoneInfo("UTC")
    start_of_day_utc = start_of_day_local.astimezone(utc_tz)
    end_of_day_utc = end_of_day_local.astimezone(utc_tz)

    today_runs_count = ScriptRun.objects.filter(
        script__in=user_scripts,
        started_at__gte=start_of_day_utc,
        started_at__lte=end_of_day_utc
    ).count()
    
    # Получаем статистику новых вакансий из последних запусков каждого скрипта
    new_vacancies_total = 0
    for script in user_scripts:
        last_run = script.runs.filter(status='completed').order_by('-started_at').first()
        if last_run:
            new_vacancies_total += last_run.new_vacancies or 0
    
    # Получаем общее количество уникальных вакансий за всю историю
    total_vacancies = Vacancy.objects.filter(
        script__in=user_scripts
    ).count()
    
    context = {
        'user_scripts': user_scripts,
        'recent_runs': recent_runs,
        'new_vacancies_total': new_vacancies_total,
        'total_vacancies': total_vacancies,
        'today_runs_count': today_runs_count,
    }
    return render(request, 'home.html', context)


@login_required
def script_list_view(request):
    """Список скриптов пользователя"""
    scripts = get_accessible_scripts_for_user(request.user, is_active=True)
    return render(request, 'scripts/list.html', {'scripts': scripts})


@login_required
def script_detail_view(request, script_id):
    """Детали скрипта и вакансий"""
    # Получаем скрипт с проверкой доступа
    accessible_scripts = get_accessible_scripts_for_user(request.user, is_active=None)
    script = get_object_or_404(accessible_scripts, id=script_id)
    
    # Дополнительная проверка доступа через метод модели
    if not script.has_access(request.user):
        from django.http import Http404
        raise Http404("У вас нет доступа к этому скрипту")
    
    # Получаем последний запуск
    last_run = script.runs.order_by('-started_at').first()
    
    # Получаем вакансии из последнего запуска
    new_vacancy_runs = []
    old_vacancy_runs = []
    
    if last_run:
        new_vacancy_runs = last_run.vacancy_runs.filter(is_new_in_run=True).select_related('vacancy').order_by('-found_at')
        old_vacancy_runs = last_run.vacancy_runs.filter(is_new_in_run=False).select_related('vacancy').order_by('-found_at')
    
    # Получаем общую статистику по всем вакансиям скрипта
    all_vacancies = script.vacancies.all()
    active_vacancies = all_vacancies.filter(is_active=True)
    
    context = {
        'script': script,
        'last_run': last_run,
        'new_vacancy_runs': new_vacancy_runs,
        'old_vacancy_runs': old_vacancy_runs,
        'total_vacancies': all_vacancies.count(),
        'active_vacancies_count': active_vacancies.count(),
    }
    return render(request, 'scripts/detail.html', context)


@login_required
def run_script_view(request, script_id):
    """Запуск скрипта"""
    if request.method == 'POST':
        # Получаем скрипт с проверкой доступа
        accessible_scripts = get_accessible_scripts_for_user(request.user, is_active=True)
        script = get_object_or_404(accessible_scripts, id=script_id)
        
        # Дополнительная проверка доступа через метод модели
        if not script.has_access(request.user):
            return JsonResponse({
                'success': False, 
                'error': 'У вас нет доступа к этому скрипту'
            })
        
        # Проверяем, нет ли активных запусков
        active_run = script.runs.filter(status='running').first()
        if active_run:
            return JsonResponse({
                'success': False, 
                'error': 'Скрипт уже выполняется'
            })
        
        # Создаем новый запуск
        script_run = ScriptRun.objects.create(
            script=script,
            started_by=request.user,
            status='running'
        )
        
        # Запускаем парсер в отдельном потоке
        def run_parser():
            parser = HHVacancyParserDjango(script_run)
            parser.run()
        
        thread = threading.Thread(target=run_parser)
        thread.daemon = True
        thread.start()
        
        return JsonResponse({
            'success': True, 
            'run_id': script_run.id,
            'message': 'Скрипт запущен'
        })
    
    return JsonResponse({'success': False, 'error': 'Неправильный метод запроса'})


@login_required
def script_status_view(request, run_id):
    """Получение статуса выполнения скрипта"""
    try:
        # Получаем доступные скрипты и фильтруем запуск
        accessible_scripts = get_accessible_scripts_for_user(request.user, is_active=None)
        script_run = ScriptRun.objects.get(
            id=run_id, 
            script__in=accessible_scripts
        )
        
        return JsonResponse({
            'success': True,
            'status': script_run.status,
            'total_found': script_run.total_found,
            'new_vacancies': script_run.new_vacancies,
            'existing_vacancies': script_run.existing_vacancies,
            'error_message': script_run.error_message,
            'completed_at': script_run.completed_at.isoformat() if script_run.completed_at else None
        })
    except ScriptRun.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Запуск не найден'})


@login_required
def script_history_view(request):
    """История запусков скриптов"""
    accessible_scripts = get_accessible_scripts_for_user(request.user, is_active=None)
    runs = ScriptRun.objects.filter(
        script__in=accessible_scripts
    ).order_by('-started_at')
    
    paginator = Paginator(runs, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'scripts/history.html', {'page_obj': page_obj})


@login_required
def vacancies_view(request, run_id):
    """Просмотр вакансий по конкретному запуску"""
    # Получаем доступные скрипты и фильтруем запуск
    accessible_scripts = get_accessible_scripts_for_user(request.user, is_active=None)
    script_run = get_object_or_404(
        ScriptRun, 
        id=run_id, 
        script__in=accessible_scripts
    )
    
    # Дополнительная проверка доступа
    if not script_run.script.has_access(request.user):
        from django.http import Http404
        raise Http404("У вас нет доступа к этому скрипту")
    
    vacancy_filter = request.GET.get('filter', 'all')
    
    if vacancy_filter == 'new':
        vacancy_runs = script_run.vacancy_runs.filter(is_new_in_run=True)
    elif vacancy_filter == 'existing':
        vacancy_runs = script_run.vacancy_runs.filter(is_new_in_run=False)
    else:
        vacancy_runs = script_run.vacancy_runs.all()
    
    vacancy_runs = vacancy_runs.select_related('vacancy').order_by('-found_at')
    
    paginator = Paginator(vacancy_runs, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'script_run': script_run,
        'page_obj': page_obj,
        'current_filter': vacancy_filter,
    }
    return render(request, 'scripts/vacancies.html', context)


@login_required
def export_vacancies_excel(request, script_id):
    """Экспорт вакансий в Excel с двумя листами"""
    # Получаем скрипт, к которому у пользователя есть доступ
    accessible_scripts = get_accessible_scripts_for_user(request.user, is_active=None)
    script = get_object_or_404(accessible_scripts, id=script_id)
    
    # Проверяем доступ дополнительно через метод модели
    if not script.has_access(request.user):
        from django.http import Http404
        raise Http404("У вас нет доступа к этому скрипту")
    
    # Получаем последний запуск скрипта
    last_run = script.runs.order_by('-started_at').first()
    if not last_run:
        messages.error(request, 'Нет данных для экспорта. Сначала запустите скрипт.')
        return redirect('scripts:detail', script_id=script.id)
    
    # Импортируем библиотеку для работы с Excel
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from django.utils import timezone
    import io
    
    # Создаем книгу Excel
    workbook = Workbook()
    
    # Удаляем дефолтный лист
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    
    # Стили для заголовков
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Стили для обычного текста
    normal_alignment = Alignment(vertical="top", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Получаем данные о вакансиях
    new_vacancy_runs = last_run.vacancy_runs.filter(is_new_in_run=True).select_related('vacancy').order_by('-found_at')
    existing_vacancy_runs = last_run.vacancy_runs.filter(is_new_in_run=False).select_related('vacancy').order_by('-found_at')
    
    # Создаем лист "Новые вакансии"
    new_sheet = workbook.create_sheet(title="Новые вакансии")
    
    # Заголовки для новых вакансий (обновленные с новыми полями)
    headers = [
        "№", "Название вакансии", "Компания", "Регион", "Зарплата", "Ссылка", 
        "Найдено по запросу", "Дата публикации", "Дата обнаружения", "Количество обнаружений"
    ]
    
    # Записываем заголовки
    for col_num, header in enumerate(headers, 1):
        cell = new_sheet.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Записываем данные новых вакансий
    for row_num, vacancy_run in enumerate(new_vacancy_runs, 2):
        vacancy = vacancy_run.vacancy
        data = [
            row_num - 1,  # Номер по порядку
            vacancy.title,
            vacancy.company,
            vacancy.area_name or "Не указан",
            vacancy.salary,
            vacancy.url,
            vacancy.found_by_query or "Не указан",
            vacancy.published_at.strftime("%d.%m.%Y %H:%M") if vacancy.published_at else "Не указана",
            vacancy_run.found_at.strftime("%d.%m.%Y %H:%M"),
            vacancy.times_found
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = new_sheet.cell(row=row_num, column=col_num, value=value)
            cell.alignment = normal_alignment
            cell.border = border
    
    # Настраиваем ширину колонок для новых вакансий (обновленные ширины)
    column_widths = [5, 50, 30, 20, 20, 60, 25, 20, 20, 15]
    for col_num, width in enumerate(column_widths, 1):
        new_sheet.column_dimensions[get_column_letter(col_num)].width = width
    
    # Создаем лист "Существующие вакансии"
    existing_sheet = workbook.create_sheet(title="Существующие вакансии")
    
    # Записываем заголовки для существующих вакансий
    for col_num, header in enumerate(headers, 1):
        cell = existing_sheet.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Записываем данные существующих вакансий
    for row_num, vacancy_run in enumerate(existing_vacancy_runs, 2):
        vacancy = vacancy_run.vacancy
        data = [
            row_num - 1,  # Номер по порядку
            vacancy.title,
            vacancy.company,
            vacancy.area_name or "Не указан",
            vacancy.salary,
            vacancy.url,
            vacancy.found_by_query or "Не указан",
            vacancy.published_at.strftime("%d.%m.%Y %H:%M") if vacancy.published_at else "Не указана",
            vacancy_run.found_at.strftime("%d.%m.%Y %H:%M"),
            vacancy.times_found
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = existing_sheet.cell(row=row_num, column=col_num, value=value)
            cell.alignment = normal_alignment
            cell.border = border
    
    # Настраиваем ширину колонок для существующих вакансий
    for col_num, width in enumerate(column_widths, 1):
        existing_sheet.column_dimensions[get_column_letter(col_num)].width = width
    
    # Добавляем информационный лист
    info_sheet = workbook.create_sheet(title="Информация", index=0)
    info_sheet.merge_cells('A1:D1')
    title_cell = info_sheet['A1']
    title_cell.value = f"Отчет по вакансиям: {script.name}"
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center")
    
    # Добавляем информацию о запуске
    info_data = [
        ["Скрипт:", script.name],
        ["Поисковые запросы:", script.get_search_summary()],
        ["Регион:", script.get_region_display_name()],
        ["Дата запуска:", last_run.started_at.strftime("%d.%m.%Y %H:%M")],
        ["Статус:", last_run.get_status_display()],
        ["Всего найдено:", last_run.total_found],
        ["Новых вакансий:", last_run.new_vacancies],
        ["Существующих вакансий:", last_run.existing_vacancies],
        ["Экспортировано:", timezone.now().strftime("%d.%m.%Y %H:%M")]
    ]
    
    # Добавляем статистику по запросам, если есть
    stats = last_run.get_queries_stats()
    if stats:
        info_data.append(["", ""])  # Пустая строка
        info_data.append(["Детальная статистика:", ""])
        for query, query_stats in stats.items():
            info_data.append([f"  '{query}':", ""])
            info_data.append([f"    Найдено в API:", query_stats.get('total_found', 0)])
            info_data.append([f"    Собрано:", query_stats.get('collected_vacancies', 0)])
            info_data.append([f"    Уникальных:", query_stats.get('unique_vacancies', 0)])
            info_data.append([f"    Новых:", query_stats.get('new_vacancies', 0)])
            info_data.append([f"    Существующих:", query_stats.get('existing_vacancies', 0)])
    
    for row_num, (label, value) in enumerate(info_data, 3):
        info_sheet[f'A{row_num}'] = label
        info_sheet[f'B{row_num}'] = value
        info_sheet[f'A{row_num}'].font = Font(bold=True)
    
    # Настраиваем ширину колонок для информационного листа
    info_sheet.column_dimensions['A'].width = 25
    info_sheet.column_dimensions['B'].width = 40
    
    # Подготавливаем ответ
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Генерируем имя файла
    filename = f"vacancies_{script.name}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Сохраняем книгу в ответ
    workbook.save(response)
    
    return response


from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required

@require_POST
@login_required
def delete_script_run_view(request, run_id):
    """
    Удаление запуска скрипта со всеми связанными данными
    """
    try:
        # Получаем доступные скрипты пользователя
        accessible_scripts = get_accessible_scripts_for_user(request.user, is_active=None)
        
        # Получаем запуск и проверяем доступ
        script_run = get_object_or_404(ScriptRun, id=run_id, script__in=accessible_scripts)
        
        # Дополнительная проверка прав доступа через метод модели
        if not script_run.script.has_access(request.user):
            return JsonResponse({
                'success': False,
                'error': 'У вас нет прав на удаление этого запуска'
            }, status=403)
        
        # Подсчитываем данные для статистики ПЕРЕД удалением
        vacancy_runs_count = VacancyRun.objects.filter(script_run=script_run).count()
        vacancy_ids = VacancyRun.objects.filter(script_run=script_run).values_list('vacancy_id', flat=True)
        unique_vacancy_ids = set(vacancy_ids)
        
        # Проверяем, какие вакансии больше не связаны с другими запусками
        vacancies_to_delete = []
        for vacancy_id in unique_vacancy_ids:
            # Проверяем, есть ли у вакансии другие запуски кроме текущего
            other_runs_count = VacancyRun.objects.filter(
                vacancy_id=vacancy_id
            ).exclude(script_run=script_run).count()
            
            if other_runs_count == 0:
                # Вакансия больше нигде не используется, можно удалить
                vacancies_to_delete.append(vacancy_id)
        
        script_name = script_run.script.name
        run_date = script_run.started_at
        
        # Удаляем данные в правильном порядке
        # 1. Удаляем VacancyRun (связи между вакансиями и запусками)
        VacancyRun.objects.filter(script_run=script_run).delete()
        
        # 2. Удаляем вакансии, которые больше не используются
        if vacancies_to_delete:
            Vacancy.objects.filter(id__in=vacancies_to_delete).delete()
        
        # 3. Удаляем сам запуск
        script_run.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Запуск "{script_name}" от {run_date.strftime("%d.%m.%Y %H:%M")} успешно удален',
            'stats': {
                'deleted_vacancy_runs': vacancy_runs_count,
                'deleted_vacancies': len(vacancies_to_delete),
                'total_unique_vacancies': len(unique_vacancy_ids)
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Ошибка при удалении: {str(e)}'
        }, status=500)
